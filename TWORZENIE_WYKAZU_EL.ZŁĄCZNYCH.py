import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Lista dopuszczalnych nazw arkuszy
acceptable_sheet_names = ["ZŁĄCZNE", "złączne", "Złączne", "zlaczne", "Zlaczne", "ZLACZNE"]

def autofit_column_widths(filename, sheet_name="Sheet1"):
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    for column_cells in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2  # Dodajemy trochę luzu

    wb.save(filename)

def main():
    # Okno dialogowe do wyboru pliku
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Wybierz plik Excel",
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")]
    )

    if not file_path:
        print("Nie wybrano pliku.")
        return

    # Wczytanie arkuszy
    xls = pd.ExcelFile(file_path)

    # Szukanie arkusza o dokładnej nazwie z listy
    for sheet_name in xls.sheet_names:
        if sheet_name in acceptable_sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            output_file = "elementy_złączne.xlsx"
            df.to_excel(output_file, index=False)

            autofit_column_widths(output_file)

            print(f"Zapisano arkusz '{sheet_name}' do '{output_file}' z dopasowaną szerokością kolumn.")
            return

    print("Nie znaleziono zakładki o nazwie z listy dopuszczalnych nazw:" +
          ", ".join(acceptable_sheet_names))

if __name__ == "__main__":
    main()
