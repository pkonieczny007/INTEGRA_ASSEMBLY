import openpyxl
import glob
import os

# Wyszukaj plik .xlsm zaczynający się od INTEGRA
xlsm_file_path = glob.glob('INTEGRA*.xlsm')[0]  # Zakładamy, że plik znajduje się w katalogu /mnt/data/

# Załaduj plik .xlsm
xlsm_file = openpyxl.load_workbook(xlsm_file_path, keep_vba=False)

# Utwórz nowy plik .xlsx
xlsx_file = openpyxl.Workbook()

# Usuń domyślny arkusz utworzony przez openpyxl w nowym pliku .xlsx
xlsx_file.remove(xlsx_file.active)

# Skopiuj wszystkie arkusze z pliku .xlsm do .xlsx
for sheet_name in xlsm_file.sheetnames:
    # Utwórz nowy arkusz o tej samej nazwie
    xlsx_file.create_sheet(sheet_name)
    source = xlsm_file[sheet_name]
    destination = xlsx_file[sheet_name]

    # Skopiuj dane
    for row in source.rows:
        for cell in row:
            destination[cell.coordinate].value = cell.value

# Zapisz nowy plik .xlsx
output_directory = '/IMPORTERY/'
os.makedirs(output_directory, exist_ok=True)  # Tworzy katalog, jeśli nie istnieje
xlsx_file_path = os.path.join(output_directory, 'wykaz.xlsx')
xlsx_file.save(xlsx_file_path)

# Zwróć pełną ścieżkę do nowego pliku
xlsx_file_path
